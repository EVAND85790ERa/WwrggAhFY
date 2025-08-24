# 代码生成时间: 2025-08-25 00:39:00
import quart
from quart import request, jsonify
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

# 定义路由和视图函数
app = quart.Quart(__name__)

@app.route('/create_excel', methods=['POST'])
def create_excel():
    # 获取前端传递的数据
    data = request.get_json()
    
    # 检查数据是否有效
    if not data or 'headers' not in data or 'rows' not in data:
        return jsonify({'error': 'Invalid data'}), 400
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    
    # 添加表头
    for col, header in enumerate(data['headers'], start=1):
        ws.cell(row=1, column=col, value=header)
    
    # 添加数据行
    for row_idx, row in enumerate(data['rows'], start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 生成文件并保存
    file_path = 'generated_excel.xlsx'
    wb.save(file_path)
    
    # 返回文件路径供前端下载
    return jsonify({'file_path': file_path}), 200

if __name__ == '__main__':
    app.run(debug=True)

"""
Excel表格自动生成器

这个程序使用Quart框架和openpyxl库来创建Excel表格。
用户通过POST请求发送数据，程序将数据转换成Excel文件，并返回文件路径供用户下载。

请求数据格式:
{
    "headers": ["Header1", "Header2", ...],
    "rows": [["Value11", "Value12", ...], ["Value21", "Value22", ...], ...]
}

响应数据格式:
{
    "file_path": "generated_excel.xlsx"
}
"""